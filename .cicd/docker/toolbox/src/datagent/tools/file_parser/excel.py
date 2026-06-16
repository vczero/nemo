import json
from openpyxl import load_workbook


def parse_excel(file_path: str) -> dict:
    """Parse Excel file, return jsonlines + statistics."""
    wb = load_workbook(file_path, data_only=True)
    jsonlines = []
    stats = {"sheet_count": len(wb.sheetnames), "sheets": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append(dict(zip(headers, row)))

        for row in rows:
            jsonlines.append(json.dumps(row, ensure_ascii=False))

        stats["sheets"][sheet_name] = {"row_count": len(rows), "col_count": len(headers)}

    return {
        "format": "jsonlines",
        "content": "\n".join(jsonlines),
        "summary": stats,
    }
